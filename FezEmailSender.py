import os
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Email, To, Content
from datetime import datetime
import json
from openpyxl import load_workbook
from string import ascii_letters

# Convert string to integer with validation
def convertToInt(msg):
    flag = False
    while flag == False:
        try:
            result = int(input(msg))
            flag = True
        except:
            print("Error! Check your number format and try again")
    return result

# String input with validation
def inputNoBlank(msg):
    flag = False
    while flag == False:
        result = input(msg)
        if len(result) > 0:
            flag = True
        else:
            print("Empty input! Please try again")
    return result

# Convert Excel column letter into number
def colLetterToNumber(col):
    print(f"col: {col}")
    num = 0
    for char in col:
        #if char in ascii_letters:
        num = num * 26 + (ord(char) - ord('A')) + 1
    return num

# Extract emails from an Excel spreadsheet
def emailExtractExcel(emailList):
    # Store for sheet index
    sheetIndex = None

    # Workbook object
    workbookObj = None

    # List of emails which could not be successfully added
    unsuccessfulContacts = []

    # Prompt user for directory for workbook, and try to load workbook
    flag = False
    while flag == False:
        try:
            workbookDir = inputNoBlank("Drag in or copy & paste the path to the Excel spreadsheet and press Enter: ")

            # Load workbook, using directory without inverted commas at beginning and end
            if workbookDir[0] == "'":
                workbookObj = load_workbook(workbookDir[1:len(workbookDir) - 1:1], read_only=True)
            else:
                workbookObj = load_workbook(workbookDir, read_only=True)
            flag = True

        except Exception as e:
            print("You did not enter a valid spreadsheet file path. Please try again.")
            print(e)

    # Ask which sheet the user wants to import
    workbookSheetNames = workbookObj.sheetnames
    print("\nThese are the sheet names currently in the spreadsheet:\n")
    for i in range(0, len(workbookSheetNames)):
        print(f"[{i + 1}] {workbookSheetNames[i]}")
    print("\nWhich sheet do you want to load?")
    flag = False
    while flag == False:
        sheetIndex = convertToInt("Enter a number: ")
        if sheetIndex > len(workbookObj.worksheets) or sheetIndex < 1:
            print("Invalid spreadsheet index. Please try again")
        else:
            flag = True

    # Loop for cell importing
    flag = False
    while flag == False:
        try:
            # Ask for cell where names begin
            cellStartName = inputNoBlank("Enter the cell letter & number where the first surname begins, e.g. A3: ").upper()

            # Ask for the colum where the first name is
            columnFirstNameLetter = inputNoBlank("Enter the column letter which contains the first names: ").upper()

            # Ask for column where email addresses are located
            columnEmailLetter = inputNoBlank("Enter the column letter which contains email addresses: ").upper()

            # Ask for last row number
            rowLastEntry = convertToInt("Enter the number of the last row: ")

            #xprint(f"Value in selected cell: {workbookObj.worksheets[sheetIndex][cellStartName].value}")

            # Convert cell letter and number into column and row
            cellStartColString = ""
            currentChar = 0
            for char in cellStartName:
                if char in ascii_letters:
                    cellStartColString = cellStartColString + char
                    currentChar += 1
                else:
                    break
            rowStart = int(cellStartName[currentChar:len(cellStartName):1])
            columnSurnameNum = colLetterToNumber(cellStartColString)
            columnFirstNameNum = colLetterToNumber(columnFirstNameLetter)
            columnEmailNum = colLetterToNumber(columnEmailLetter)

            # Extract information from spreadsheet, looping until we reach the last row
            currentSurname = None
            currentFirstName = None
            currentEmail = None
            for currentRow in range(rowStart, rowLastEntry + 1):

                print(f"Current cells: first name: ({currentRow}, {columnFirstNameNum}), surname: ({currentRow}, {columnSurnameNum}), address: ({currentRow}, {columnEmailNum})")
                
                # Check if the current surname cell has an entry, then populate the other entries from there
                currentSurname = workbookObj.worksheets[sheetIndex].cell(row=currentRow, column=columnSurnameNum).value
                if currentSurname != None:
                    currentFirstName = workbookObj.worksheets[sheetIndex].cell(row=currentRow, column=columnFirstNameNum).value
                    if currentFirstName != None:
                        currentEmail = workbookObj.worksheets[sheetIndex].cell(row=currentRow, column=columnEmailNum).value

                        # ***** CONTINUE FROM HERE *****
                        # Email column not being retrieved correctly
                        
                        # If the contact has no email address, add their details to the unsuccessful contacts list, otherwise add to the email list
                        if currentEmail == None:
                            unsuccessfulContacts.append(f"{currentFirstName} {currentSurname}")
                        else:
                            emailList.append({'address': currentEmail, 'firstName': currentFirstName, 'surname': currentSurname})
                            print(f"==== Added {currentFirstName} {currentSurname} with address {currentEmail} =====")

            flag = True

        except Exception as e:
            print("Error! Please try again. You may need to check the contents of your spreadsheet. See details below")
            print(e)

    # List contacts with missing email addresses if there are any
    if len(unsuccessfulContacts) > 0:
        print("\nThe following contacts do not have email addresses:")
        for contact in unsuccessfulContacts:
            print(contact)

# Extract emails from the user manually
def emailExtractManual(emailList):
    # Enter list of email addresses
    flag = True
    while flag == True:
        emailList.append({'address': inputNoBlank("Enter a recipient email address: "), 'firstName': inputNoBlank("Enter a first name: "), 'surname': inputNoBlank("Enter a surname: ")})
        check = input("Stop adding? (type y to end, or any other key to add another address)")
        if check == 'y':
            flag = False

# Main program
def mainFunc():
        # Function index mapping
    funcMap = {'1': emailExtractExcel, '2': emailExtractManual}

    # Lists
    emailList = []
    paragraphs = []

    # Welcome message
    print("===== Welcome to the FezEmailSender =====\n")

    # Create an emails directory if there isn't one already
    try:
        os.mkdir('./Emails')
    except:
        print('===== The directory Emails already exists. Emails will be stored here in .json format =====\n')
    else:
        print('===== The directory Emails has been created in the folder where this script is located ====')
        print('===== It contains all of the emails which you have sent in .json format =====\n')

    # Enter the sender details and pass them to the helpers
    fromEmail = Email(inputNoBlank("Enter the email address to send from: "))
    emailSubject = inputNoBlank("Enter the email subject for all of the recipients: ")
    yourName = inputNoBlank("Enter your full name: ")

    # Let user choose which mode to load program in
    choiceFlag = False
    while choiceFlag == False:
        emailExtractChoice = inputNoBlank("How do you want to import the addresses? 1. From an Excel spreadsheet, or 2. Manually. Enter number: ")
        try:
            funcMap[emailExtractChoice](emailList)
            choiceFlag = True
        except Exception as e:
            print("You did not enter a valid number. Please try again")
            print(e)

    # Construct message
    numberParagraphs = convertToInt("\nEnter the number of paragraphs to use: ")
    for currentIndex in range(0, numberParagraphs):
        paragraphs.append(inputNoBlank(f"Enter paragraph {currentIndex + 1}:\n"))
    lastSentence = inputNoBlank("Enter your last sentence:\n")
    endGreeting = inputNoBlank("Enter your end greeting:\n")

    # Send email for each recipient
    for currentPerson in emailList:
        stillWriting = True
        while stillWriting == True:

            # Writing message
            print(f"\n===== Now emailing {currentPerson['firstName']} {currentPerson['surname']} =====\n")
            messageText = f"<p>Hello {currentPerson['firstName']}</p>"
            for currentParagraph in paragraphs:
                print(currentParagraph)
                messageText = messageText + "<p>" + currentParagraph + " " + inputNoBlank("Continue this paragraph:\n") + "</p>"
                print()
            messageText = messageText + '<p>' + lastSentence + '</p><p>' + endGreeting + "</p><p>" + yourName + '</p>'
            print(f"\nFull message for {currentPerson['firstName']} {currentPerson['surname']}:\n\n***")
            messageTextToDisplay = messageText.replace('</p>', '</p>\n\n')
            print(messageTextToDisplay)
            print('***\n')
            print(f"Recipient email address: {currentPerson['address']}")
            check = input("\nAre you sure you want to send this message? This action cannot be undone (type y)")

            # Confirm from the user that they want to send the message
            if check == 'y':
                stillWriting = False

                # Pass the to address to the helper
                toEmail = To(currentPerson['address'])
                content = Content('text/html', messageText)

                # Construct the message
                message = Mail(fromEmail, toEmail, emailSubject, content)

                # Save a .json representation of the message
                mailJSON = message.get()

                # Add sending and receiving names to the message
                mailJSON['from']['name'] = yourName
                mailJSON['personalizations'][0]['to'][0]['name'] = currentPerson['firstName'] + " " + currentPerson['surname']

                # Save .json file to disk
                timestamp = datetime.now()
                fileName = f"{timestamp}".replace(" ", "_") + "_" + currentPerson['firstName'] + "-" + currentPerson['surname']
                fileName.replace('/', '-')
                fileName.replace(' ', '-')
                with open(f"./Emails/{fileName}.json", 'w') as outfile:
                    json.dump(mailJSON, outfile)

                # Try to send the message
                responseFlag = False
                while responseFlag == False:
                    try:
                        # Use SendGrid API linked to the OS
                        sg = SendGridAPIClient(api_key=os.environ.get('SENDGRID_API_KEY'))
                        response = sg.client.mail.send.post(request_body=mailJSON)
                        print(f"\nSendGrid response for {currentPerson['address']}")
                        print(response.status_code)
                        print(response.body)
                        print(response.headers)
                        responseFlag = True
                    except Exception as e:
                        print(e)
                        retryCheck = input("\nTry again? Type y to try again. Press any other key to move onto the next email (note this faulty email down)")
                        if retryCheck != "y":
                            responseFlag = True
            
            # If user does not like the message, they can enter it again
            else:
                print("===== Enter the response for your message again below =====\n")

    print("\n==== Operation complete =====")

if __name__ == "__main__":
    mainFunc()